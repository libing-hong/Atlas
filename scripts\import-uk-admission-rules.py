"""Import the confidential UK workbook into a private, query-only index.

The workbook itself is never copied into public/, the client bundle, or an API response.
The generated index intentionally keeps only sheet-level matching metadata; ambiguous
matches are returned for manual review instead of being treated as eligible.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl


def normalize(value: str) -> str:
    return re.sub(r"[\s\u3000,，.。'’\"“”()（）-]", "", value).lower()


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: import-uk-admission-rules.py <workbook.xlsx>")
    workbook_path = Path(sys.argv[1]).resolve()
    output_path = Path(__file__).resolve().parents[1] / ".atlas-private" / "uk-admission-rules.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheets = []
    for sheet_name in workbook.sheetnames[1:]:
        sheet = workbook[sheet_name]
        values = [str(cell).strip() for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None and str(cell).strip()]
        search_text = normalize(" ".join(values))
        joined = " ".join(values)
        percentages = sorted({int(value) for value in re.findall(r"(?<!\d)(\d{2})\s*%", joined)})
        if any(token in joined.lower() for token in ("case by case", "case-by-case")):
            rule_type = "case_by_case"
        elif any(token in joined for token in ("985", "211", "双一流")):
            rule_type = "985_211_double_first_class"
        elif "Tier" in joined or "tier" in joined:
            rule_type = "tier_list"
        elif percentages:
            rule_type = "program_specific"
        else:
            rule_type = "not_supported"
        sheets.append({
            "sheetName": sheet_name,
            "sourceYear": 2026,
            "ruleType": rule_type,
            "percentages": percentages[:20],
            "searchText": search_text,
            "confidential": True,
            "importedAt": date.today().isoformat(),
        })
    output_path.write_text(json.dumps({"version": "2026.07", "sourceType": "private_partner_workbook", "sheets": sheets}, ensure_ascii=False), encoding="utf-8")
    print(f"Imported {len(sheets)} UK institution sheets into {output_path}")


if __name__ == "__main__":
    main()
